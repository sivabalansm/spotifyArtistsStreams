#!/usr/bin/env python3
"""
Spotify for Artists - Playlist Streams Scraper

Scrapes playlist stream data from artists.spotify.com and fills in an xlsx spreadsheet.

Usage:
    # First time: login and scrape
    python3 scraper.py --login --xlsx "Playlist performance Overview March.xlsx"

    # Subsequent runs (reuses saved session)
    python3 scraper.py --xlsx "Playlist performance Overview March.xlsx"

    # Standalone mode (no xlsx, just scrape URLs)
    python3 scraper.py <url_or_id> [url_or_id ...]
    python3 scraper.py --file songs.txt
"""

import argparse
import copy
import csv
import json
import os
import re
import shutil
import sys
import time
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# Playlist name mapping: xlsx column header -> name as it appears on Spotify for Artists
PLAYLIST_MAP = {
    "Mainstage": "Tomorrowland MainStage 2026",
    "Essentials": "Tomorrowland Essentials",
    "Guest List": "Tomorrowland Guest List",
    "Soave": "Tomorrowland 2026 Playlist",  # partial match — emoji varies
}

# Column layout in xlsx (0-indexed): each playlist has 3 columns (7d, 28d, 12mo)
# C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10, K=11, L=12, M=13, N=14
PLAYLIST_COLUMNS = {
    "Mainstage": {"7 days": "C", "28 days": "D", "12 months": "E"},
    "Essentials": {"7 days": "F", "28 days": "G", "12 months": "H"},
    "Guest List": {"7 days": "I", "28 days": "J", "12 months": "K"},
    "Soave": {"7 days": "L", "28 days": "M", "12 months": "N"},
}

TIME_PERIOD_LABELS = ["7 days", "28 days", "12 months"]

AUTH_STATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".auth_state.json")
OUTPUT_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "playlist_streams.csv")


def parse_song_input(song_input: str) -> dict:
    """Extract track ID and optionally artist ID from a Spotify URL or plain ID."""
    song_input = song_input.strip()

    match = re.search(
        r"artists\.spotify\.com/c/artist/([a-zA-Z0-9]+)/song/([a-zA-Z0-9]+)",
        song_input,
    )
    if match:
        return {"artist_id": match.group(1), "track_id": match.group(2)}

    match = re.search(r"track/([a-zA-Z0-9]{22})", song_input)
    if match:
        return {"track_id": match.group(1)}

    if re.match(r"^[a-zA-Z0-9]{22}$", song_input):
        return {"track_id": song_input}

    raise ValueError(f"Cannot extract track ID from: {song_input}")


def login_and_save_state(playwright):
    """Open browser for manual login, save auth state for reuse."""
    print("\n🔐 Opening browser for Spotify for Artists login...")
    print("   Log in manually, then press ENTER here when you're on the dashboard.\n")

    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto("https://artists.spotify.com/c/artist/")

    input("   >> Press ENTER after you've logged in and see the dashboard... ")

    context.storage_state(path=AUTH_STATE_PATH)
    print("   Auth state saved. Future runs will reuse this session.\n")
    browser.close()


def detect_artist_id(page) -> str:
    """Extract the artist ID from the current URL after login."""
    url = page.url
    match = re.search(r"/artist/([a-zA-Z0-9]{22})", url)
    if match:
        return match.group(1)
    page.goto("https://artists.spotify.com/c/artist/", wait_until="networkidle", timeout=30000)
    time.sleep(5)
    url = page.url
    match = re.search(r"/artist/([a-zA-Z0-9]{22})", url)
    if match:
        return match.group(1)
    # Try clicking through any intermediate page
    try:
        links = page.query_selector_all("a[href*='/artist/']")
        for link in links:
            href = link.get_attribute("href") or ""
            m = re.search(r"/artist/([a-zA-Z0-9]{22})", href)
            if m:
                return m.group(1)
    except Exception:
        pass
    raise RuntimeError("Could not detect artist ID from URL. Use --artist flag to set it manually.")


def switch_to_artist(page, artist_name):
    """Search for an artist in the top 'Search artists' bar and switch to them.

    Returns the artist_id from the URL after switching, or None if not found.
    """
    print(f"   Switching to artist: {artist_name}...")

    # Click the "Search artists" textbox at the top
    search_box = page.query_selector("input[aria-label='Search artists'], input[placeholder='Search artists']")
    if not search_box:
        print("   ⚠ Could not find artist search bar")
        return None

    search_box.fill("")
    time.sleep(0.5)
    search_box.fill(artist_name)
    time.sleep(2)  # wait for search results dropdown

    # Click the matching artist from the dropdown results
    # Results appear as links with /artist/{id}/home in the href
    results = page.query_selector_all("a[href*='/artist/'][href*='/home']")
    for result in results:
        result_text = result.inner_text().strip()
        if artist_name.lower() in result_text.lower():
            href = result.get_attribute("href") or ""
            result.click()
            time.sleep(3)
            m = re.search(r"/artist/([a-zA-Z0-9]+)", page.url)
            if m:
                print(f"   ✓ Switched to {artist_name} (ID: {m.group(1)})")
                return m.group(1)

    # Fallback: press Enter and see if it navigates
    search_box.press("Enter")
    time.sleep(3)
    m = re.search(r"/artist/([a-zA-Z0-9]+)", page.url)
    if m:
        return m.group(1)

    print(f"   ✗ Could not find artist '{artist_name}'")
    return None


def dismiss_popups(page):
    """Dismiss any modal dialogs that block interaction."""
    try:
        close_btn = page.query_selector("[data-testid='try-it-out-sos-modal'] button:has-text('Close')")
        if close_btn and close_btn.is_visible():
            close_btn.click()
            time.sleep(0.5)
    except Exception:
        pass
    try:
        cookie_btn = page.query_selector("dialog button:has-text('Close')")
        if cookie_btn and cookie_btn.is_visible():
            cookie_btn.click()
            time.sleep(0.5)
    except Exception:
        pass


def find_song_by_search(page, artist_id, artist_name, song_title):
    """Search for a song in the Music > Songs tab and return (track_id, artist_id).

    Flow:
    1. Search for artist in top search bar and switch to them
    2. Navigate to Music > Songs
    3. Use the song search box to filter
    4. Find the matching row and extract the track ID

    Returns (track_id, artist_id) or (None, artist_id) if not found.
    """
    # Step 1: Always switch to the correct artist by name
    # (different songs may be by different artists you manage)
    new_aid = switch_to_artist(page, artist_name)
    if new_aid:
        artist_id = new_aid

    # Step 2: Navigate to Music > Songs
    songs_url = f"https://artists.spotify.com/c/artist/{artist_id}/music/songs"
    print(f"   Navigating to songs catalog...")
    page.goto(songs_url, wait_until="networkidle", timeout=30000)
    time.sleep(3)

    dismiss_popups(page)

    # Step 3: Use the search box to filter songs
    search_input = page.query_selector("input[aria-label='Search'], input[placeholder*='Search songs']")
    if not search_input:
        # Try broader selector
        search_inputs = page.query_selector_all("input[type='text'], input[type='search']")
        for inp in search_inputs:
            placeholder = inp.get_attribute("placeholder") or ""
            aria = inp.get_attribute("aria-label") or ""
            if "search" in placeholder.lower() or "search" in aria.lower():
                if "artist" not in placeholder.lower():  # skip the top artist search bar
                    search_input = inp
                    break

    if search_input:
        search_input.fill("")
        time.sleep(0.5)
        search_input.fill(song_title)
        time.sleep(3)  # wait for table to filter
    else:
        print("   ⚠ No song search input found, scanning full table...")

    # Step 4: Find matching song row
    rows = page.query_selector_all("tbody tr")

    # Exact match first
    for row in rows:
        heading = row.query_selector("h3")
        if heading:
            heading_text = heading.inner_text().strip()
            if song_title.lower() == heading_text.lower():
                link = row.query_selector("a[href*='/song/']")
                if link:
                    href = link.get_attribute("href") or ""
                    m = re.search(r"/song/([a-zA-Z0-9]+)", href)
                    if m:
                        print(f"   ✓ Found track ID: {m.group(1)}")
                        return (m.group(1), artist_id)
                # Click row to navigate
                row.click()
                time.sleep(2)
                m = re.search(r"/song/([a-zA-Z0-9]+)", page.url)
                if m:
                    print(f"   ✓ Found track ID: {m.group(1)}")
                    return (m.group(1), artist_id)

    # Partial match fallback
    for row in rows:
        heading = row.query_selector("h3")
        if heading:
            heading_text = heading.inner_text().strip()
            if song_title.lower() in heading_text.lower():
                link = row.query_selector("a[href*='/song/']")
                if link:
                    href = link.get_attribute("href") or ""
                    m = re.search(r"/song/([a-zA-Z0-9]+)", href)
                    if m:
                        print(f"   ✓ Found track ID (partial: '{heading_text}'): {m.group(1)}")
                        return (m.group(1), artist_id)
                row.click()
                time.sleep(2)
                m = re.search(r"/song/([a-zA-Z0-9]+)", page.url)
                if m:
                    print(f"   ✓ Found track ID (partial: '{heading_text}'): {m.group(1)}")
                    return (m.group(1), artist_id)

    print(f"   ✗ Could not find '{song_title}' in catalog")
    return (None, artist_id)


def read_playlist_table(page):
    """Read the playlist table on the current page.

    Returns dict: { "playlist name (lowercase)": streams_int, ... }
    The table has columns: #, Playlist Name (h3), Made by, Streams, Date Added.
    """
    playlists = {}
    rows = page.query_selector_all("tbody tr")

    for row in rows:
        heading = row.query_selector("h3")
        if not heading:
            continue
        name = heading.inner_text().strip()

        # The Streams column is the 4th cell (index 3)
        cells = row.query_selector_all("td")
        if len(cells) >= 4:
            streams_text = cells[3].inner_text().strip().replace(",", "")
            try:
                streams = int(streams_text)
            except ValueError:
                streams = 0
            playlists[name.lower()] = streams

    return playlists


def select_time_period(page, period_label):
    """Switch the time period dropdown on the playlists page.

    period_label: "Last 7 days", "Last 28 days", or "Last 12 months"
    """
    # Find and click the date range dropdown button
    dropdown = page.query_selector("[data-testid='date-range-dropdown'] button")
    if not dropdown:
        # Fallback: find button with current period text
        dropdown = page.query_selector("button:has-text('Last 7 days'), button:has-text('Last 28 days'), button:has-text('Last 12 months')")

    if not dropdown:
        print(f"   ⚠ Could not find time period dropdown")
        return False

    dropdown.click()
    time.sleep(1)

    # Click the matching option from the listbox
    option = page.query_selector(f"[role='option']:has-text('{period_label}')")
    if not option:
        # Try alternative selectors
        options = page.query_selector_all("[role='option']")
        for opt in options:
            if period_label.lower() in opt.inner_text().strip().lower():
                option = opt
                break

    if option:
        option.click()
        time.sleep(3)  # wait for table to reload
        return True
    else:
        # Close dropdown by pressing Escape
        page.keyboard.press("Escape")
        print(f"   ⚠ Could not find option '{period_label}'")
        return False


def scrape_song_playlists(page, track_id, artist_id):
    """Navigate to a song's playlist page and extract stream data per playlist per time period.

    Flow:
    1. Go to /song/{tid}/playlists
    2. For each time period (7d, 28d, 12mo):
       a. Switch the dropdown
       b. Read the playlist table
       c. Match target playlists and store streams

    Returns dict with playlists data per xlsx column name.
    """
    url = f"https://artists.spotify.com/c/artist/{artist_id}/song/{track_id}/playlists"
    print(f"   Navigating to playlist stats for track {track_id}...")
    page.goto(url, wait_until="networkidle", timeout=30000)
    time.sleep(4)

    dismiss_popups(page)

    # Get the song title
    song_title = "Unknown"
    try:
        title_el = page.query_selector("h1")
        if title_el:
            song_title = title_el.inner_text().strip()
    except Exception:
        pass

    result = {
        "track_id": track_id,
        "song_title": song_title,
        "playlists": {},
    }

    # Initialize all target playlists
    for xlsx_name in PLAYLIST_MAP:
        result["playlists"][xlsx_name] = {"7 days": None, "28 days": None, "12 months": None}

    # Map dropdown labels to our internal period names
    period_dropdown_map = {
        "7 days": "Last 7 days",
        "28 days": "Last 28 days",
        "12 months": "Last 12 months",
    }

    for period_name, dropdown_label in period_dropdown_map.items():
        print(f"   Reading {period_name} data...")
        if not select_time_period(page, dropdown_label):
            continue

        # Read the playlist table
        table_data = read_playlist_table(page)

        # Match against our target playlists
        for xlsx_name, spotify_name in PLAYLIST_MAP.items():
            for table_playlist_name, streams in table_data.items():
                if spotify_name.lower() in table_playlist_name.lower():
                    result["playlists"][xlsx_name][period_name] = streams
                    print(f"   ✓ {xlsx_name} ({period_name}): {streams:,}")
                    break

    # Scroll down to find more playlists if any were missed
    unfound_periods = []
    for xlsx_name in PLAYLIST_MAP:
        for period_name in TIME_PERIOD_LABELS:
            if result["playlists"][xlsx_name][period_name] is None:
                unfound_periods.append(f"{xlsx_name}/{period_name}")
    if unfound_periods:
        print(f"   Not found: {unfound_periods}")

    return result


# ─── XLSX Mode ────────────────────────────────────────────────────────────────

def read_xlsx_songs(xlsx_path: str, sheet_name: str) -> list[dict]:
    """Read songs from xlsx and identify which cells need filling.

    Returns list of dicts:
    [
        {
            "row": 3,
            "artist": "John Summit",
            "song": "Silence",
            "empty_cells": {"Mainstage": ["7 days"], "Essentials": ["7 days", "28 days"], ...}
        },
        ...
    ]
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    songs = []
    for row_num in range(3, ws.max_row + 1):
        artist = ws[f"A{row_num}"].value
        song = ws[f"B{row_num}"].value

        if not artist or not song:
            continue

        # Check which cells are empty for this row
        empty_cells = {}
        has_empty = False
        for playlist_name, cols in PLAYLIST_COLUMNS.items():
            empty_periods = []
            for period, col_letter in cols.items():
                cell_val = ws[f"{col_letter}{row_num}"].value
                if cell_val is None:
                    empty_periods.append(period)
                    has_empty = True
            if empty_periods:
                empty_cells[playlist_name] = empty_periods

        if has_empty:
            songs.append({
                "row": row_num,
                "artist": str(artist).strip(),
                "song": str(song).strip(),
                "empty_cells": empty_cells,
            })

    wb.close()
    return songs


def write_one_result(xlsx_path, sheet_name, result):
    """Write a single song's scraped results into the xlsx, only filling empty cells.

    Called after each song so progress is saved incrementally.
    result dict has: row, playlists, empty_cells
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    filled_count = 0
    row_num = result["row"]
    playlists = result.get("playlists", {})
    empty_cells = result.get("empty_cells", {})

    for playlist_name, empty_periods in empty_cells.items():
        playlist_data = playlists.get(playlist_name, {})
        cols = PLAYLIST_COLUMNS[playlist_name]

        for period in empty_periods:
            value = playlist_data.get(period)
            if value is not None:
                col_letter = cols[period]
                ws[f"{col_letter}{row_num}"] = value
                filled_count += 1

    wb.save(xlsx_path)
    wb.close()
    if filled_count > 0:
        print(f"   💾 Saved {filled_count} cells to xlsx")


def run_xlsx_mode(args):
    """Main flow for xlsx mode: read songs, scrape, write back."""
    import openpyxl

    xlsx_path = args.xlsx
    sheet_name = args.sheet

    print(f"\n📊 Reading songs from '{xlsx_path}' (sheet: '{sheet_name}')...")
    songs = read_xlsx_songs(xlsx_path, sheet_name)
    print(f"   Found {len(songs)} songs with empty cells to fill\n")

    if not songs:
        print("   Nothing to do — all cells are filled!")
        return

    # Show what needs filling
    for s in songs:
        empty_summary = ", ".join(
            f"{pl}({'/'.join(p[:2] for p in periods)})"
            for pl, periods in s["empty_cells"].items()
        )
        print(f"   Row {s['row']}: {s['artist']} - {s['song']} → needs: {empty_summary}")

    print()

    # Backup original before any writes
    backup_path = xlsx_path + ".backup"
    if not os.path.exists(backup_path):
        shutil.copy2(xlsx_path, backup_path)
        print(f"   Backup saved to {backup_path}")

    with sync_playwright() as pw:
        if args.login or not os.path.exists(AUTH_STATE_PATH):
            login_and_save_state(pw)

        print("🚀 Launching browser with saved session...")
        browser = pw.chromium.launch(headless=False)
        context = browser.new_context(storage_state=AUTH_STATE_PATH)
        page = context.new_page()

        # Artist ID is resolved per-song via switch_to_artist
        artist_id = "unknown"
        total_filled = 0

        for i, song_info in enumerate(songs, 1):
            print(f"{'='*60}")
            print(f"[{i}/{len(songs)}] {song_info['artist']} - {song_info['song']}")

            # Find the track by searching the catalog
            track_id, resolved_aid = find_song_by_search(
                page, artist_id, song_info["artist"], song_info["song"]
            )

            if not track_id:
                print(f"   ❌ Skipping — song not found in catalog")
                continue

            # Scrape playlist data
            try:
                scrape_result = scrape_song_playlists(page, track_id, resolved_aid)
                result = {
                    "row": song_info["row"],
                    "playlists": scrape_result["playlists"],
                    "empty_cells": song_info["empty_cells"],
                }
                # Save to xlsx immediately after each song
                write_one_result(xlsx_path, sheet_name, result)
            except Exception as e:
                print(f"   ❌ Error scraping: {e}")

        browser.close()
        print(f"\n{'='*60}")
        print(f"✅ Done!")


# ─── Standalone Mode (original behavior) ─────────────────────────────────────

def run_standalone_mode(args):
    """Original scraper behavior: scrape URLs/IDs and output to CSV."""
    song_inputs = list(args.songs)
    if args.file:
        with open(args.file) as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    song_inputs.append(line)

    parsed_songs = []
    for s in song_inputs:
        try:
            parsed_songs.append(parse_song_input(s))
        except ValueError as e:
            print(f"⚠ Skipping: {e}")

    with sync_playwright() as pw:
        if args.login or not os.path.exists(AUTH_STATE_PATH):
            login_and_save_state(pw)

        if not parsed_songs:
            print("No tracks to process.")
            sys.exit(0)

        print("🚀 Launching browser with saved session...")
        browser = pw.chromium.launch(headless=False)
        context = browser.new_context(storage_state=AUTH_STATE_PATH)
        page = context.new_page()

        default_artist_id = None
        for p in parsed_songs:
            if "artist_id" in p:
                default_artist_id = p["artist_id"]
                break
        if not default_artist_id:
            default_artist_id = detect_artist_id(page)
        print(f"   Artist ID: {default_artist_id}")

        all_results = []
        for i, song in enumerate(parsed_songs, 1):
            track_id = song["track_id"]
            artist_id = song.get("artist_id", default_artist_id)
            print(f"\n{'='*60}")
            print(f"[{i}/{len(parsed_songs)}] Processing track: {track_id}")
            try:
                result = scrape_song_playlists(page, track_id, artist_id)
                all_results.append(result)
            except Exception as e:
                print(f"   ❌ Error: {e}")
                all_results.append({
                    "track_id": track_id,
                    "song_title": "ERROR",
                    "playlists": {},
                })

        browser.close()

        # Write CSV
        print(f"\n{'='*60}")
        print(f"📄 Writing results to {args.output}")

        target_playlists = list(PLAYLIST_MAP.values())
        fieldnames = ["track_id", "song_title"]
        for pl in target_playlists:
            for period in TIME_PERIOD_LABELS:
                fieldnames.append(f"{pl} ({period})")

        with open(args.output, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for r in all_results:
                row = {"track_id": r["track_id"], "song_title": r["song_title"]}
                for xlsx_name, spotify_name in PLAYLIST_MAP.items():
                    pl_data = r.get("playlists", {}).get(xlsx_name, {})
                    for period in TIME_PERIOD_LABELS:
                        row[f"{spotify_name} ({period})"] = pl_data.get(period)
                writer.writerow(row)

        print(f"\n✅ Done! Results saved to {args.output}")


def main():
    parser = argparse.ArgumentParser(description="Scrape Spotify for Artists playlist streams")
    parser.add_argument("songs", nargs="*", help="Spotify track URLs or IDs (standalone mode)")
    parser.add_argument("--file", "-f", help="Text file with one song URL/ID per line")
    parser.add_argument("--xlsx", help="Path to xlsx file to read songs from and write results to")
    parser.add_argument("--sheet", default="March 2026", help="Sheet name in xlsx (default: 'March 2026')")
    parser.add_argument("--login", action="store_true", help="Force re-login")
    parser.add_argument("--output", "-o", default=OUTPUT_CSV, help="Output CSV path (standalone mode)")
    args = parser.parse_args()

    if args.xlsx:
        run_xlsx_mode(args)
    elif args.songs or args.file:
        run_standalone_mode(args)
    elif args.login:
        with sync_playwright() as pw:
            login_and_save_state(pw)
    else:
        parser.print_help()
        print("\nExamples:")
        print('  python3 scraper.py --xlsx "Playlist performance Overview March.xlsx"')
        print("  python3 scraper.py https://artists.spotify.com/c/artist/.../song/.../playlists")
        print("  python3 scraper.py --file songs.txt")
        sys.exit(1)


if __name__ == "__main__":
    main()
